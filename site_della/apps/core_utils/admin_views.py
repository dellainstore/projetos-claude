"""
Views de relatório para o Django Admin da Della Instore.
Acessível em: /painel/relatorio/
"""
from decimal import Decimal

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg, Min
from django.utils import timezone


@staff_member_required
def relatorio(request):
    from apps.pedidos.models import Pedido, ItemPedido
    from apps.produtos.models import Produto, Variacao, Avaliacao
    from apps.usuarios.models import Cliente

    agora = timezone.now()
    inicio_30d = agora - timezone.timedelta(days=30)

    # Faturamento e pedidos confirmados dos últimos 30 dias
    qs_30d = Pedido.objects.filter(
        criado_em__gte=inicio_30d,
        status__in=('pagamento_confirmado', 'em_separacao', 'enviado', 'entregue'),
    )
    faturamento_30d = qs_30d.aggregate(total=Sum('total'))['total'] or Decimal('0')
    pedidos_confirmados = qs_30d.count()

    # Pedidos aguardando pagamento (todos os tempos)
    pedidos_aguardando = Pedido.objects.filter(status='aguardando_pagamento').count()

    # Totais gerais
    total_clientes = Cliente.objects.filter(is_active=True, is_staff=False).count()
    total_produtos = Produto.objects.filter(ativo=True).count()
    sem_estoque = Variacao.objects.filter(ativa=True, estoque=0).count()
    avaliacoes_pendentes = Avaliacao.objects.filter(aprovada=False).count()

    # Pedidos por status
    STATUS_LABELS = dict(Pedido.STATUS)
    pedidos_por_status = []
    for status_key, label in Pedido.STATUS:
        agg = Pedido.objects.filter(status=status_key).aggregate(
            quantidade=Count('id'), total=Sum('total')
        )
        if agg['quantidade']:
            pedidos_por_status.append({
                'label': label,
                'quantidade': agg['quantidade'],
                'total': _fmt(agg['total'] or Decimal('0')),
            })

    # Top 10 produtos mais vendidos
    top_produtos = (
        ItemPedido.objects
        .filter(pedido__status__in=('pagamento_confirmado', 'em_separacao', 'enviado', 'entregue'))
        .values('produto__id', 'produto__nome')
        .annotate(total_qty=Sum('quantidade'), receita=Sum('subtotal'))
        .order_by('-total_qty')[:10]
    )
    top_produtos = [
        {**p, 'receita': _fmt(p['receita'] or Decimal('0'))}
        for p in top_produtos
    ]

    # Estoque crítico (variações ativas com 0 a 4 unidades)
    estoque_critico = (
        Variacao.objects
        .filter(ativa=True, estoque__lte=4)
        .select_related('produto')
        .order_by('estoque', 'produto__nome')[:50]
    )

    context = {
        'title': 'Relatório Geral',
        'faturamento_30d': _fmt(faturamento_30d),
        'pedidos_confirmados': pedidos_confirmados,
        'pedidos_aguardando': pedidos_aguardando,
        'total_clientes': total_clientes,
        'total_produtos': total_produtos,
        'sem_estoque': sem_estoque,
        'avaliacoes_pendentes': avaliacoes_pendentes,
        'pedidos_por_status': pedidos_por_status,
        'top_produtos': top_produtos,
        'estoque_critico': estoque_critico,
    }
    return render(request, 'admin/relatorio.html', context)


@staff_member_required
def dashboard_pedidos(request):
    from apps.pedidos.models import Pedido, ItemPedido, HistoricoPedido

    dias_opcoes = [1, 7, 15, 30, 60, 90]
    try:
        dias = int(request.GET.get('dias', 7))
    except (TypeError, ValueError):
        dias = 7
    if dias not in dias_opcoes:
        dias = 7

    agora = timezone.now()
    inicio_periodo = agora - timezone.timedelta(days=dias)
    status_validos = ('pagamento_confirmado', 'em_separacao', 'enviado', 'entregue')

    pedidos_a_enviar = Pedido.objects.filter(
        status__in=('pagamento_confirmado', 'em_separacao'),
    ).count()

    pedidos_periodo = Pedido.objects.filter(
        criado_em__gte=inicio_periodo,
        status__in=status_validos,
    )

    historico_enviados = (
        HistoricoPedido.objects
        .filter(criado_em__gte=inicio_periodo, status_novo__in=('enviado', 'entregue'))
        .values('pedido_id')
        .distinct()
    )

    faturamento_periodo = pedidos_periodo.aggregate(total=Sum('total'))['total'] or Decimal('0')
    clientes_periodo = (
        pedidos_periodo.exclude(email='')
        .values('email')
        .distinct()
        .count()
    )
    sku_vendidos_periodo = (
        ItemPedido.objects
        .filter(pedido__in=pedidos_periodo)
        .aggregate(total=Sum('quantidade'))['total'] or 0
    )

    enviados_recentes = (
        Pedido.objects
        .filter(id__in=historico_enviados.values('pedido_id'))
        .order_by('-atualizado_em')[:10]
    )

    pedidos_para_envio = (
        Pedido.objects
        .filter(status__in=('pagamento_confirmado', 'em_separacao'))
        .order_by('criado_em')[:10]
    )

    context = {
        'title': 'Dashboard de Pedidos',
        'dias': dias,
        'dias_opcoes': dias_opcoes,
        'pedidos_a_enviar': pedidos_a_enviar,
        'pedidos_enviados': historico_enviados.count(),
        'faturamento_periodo': _fmt(faturamento_periodo),
        'clientes_periodo': clientes_periodo,
        'sku_vendidos_periodo': sku_vendidos_periodo,
        'pedidos_para_envio': pedidos_para_envio,
        'enviados_recentes': enviados_recentes,
    }
    return render(request, 'admin/dashboard_pedidos.html', context)


@staff_member_required
def instagram_refresh(request):
    """
    Limpa o cache do Instagram e força nova busca na próxima visita à homepage.
    Acessível em: /painel/instagram/refresh/
    """
    from apps.produtos.instagram import limpar_cache_instagram, buscar_posts_instagram
    from django.contrib import messages
    from django.shortcuts import redirect

    limpar_cache_instagram()

    # Tenta buscar imediatamente para validar o token
    posts = buscar_posts_instagram(limit=9)

    if posts:
        messages.success(request, f'Feed do Instagram atualizado: {len(posts)} post(s) carregados.')
    else:
        messages.warning(
            request,
            'Cache limpo, mas nenhum post foi carregado. '
            'Verifique se INSTAGRAM_ACCESS_TOKEN está configurado e válido no .env.'
        )

    return redirect('/painel/')


def _fmt(valor):
    """Formata Decimal no padrão brasileiro: 1.234,56"""
    return f'{valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


# ── Helpers de período ─────────────────────────────────────────────────────────

def _parse_periodo(request):
    from datetime import date, timedelta
    hoje = timezone.localdate()
    periodo = request.GET.get('periodo', '30d')
    data_inicio_raw = request.GET.get('data_inicio', '')
    data_fim_raw = request.GET.get('data_fim', '')

    if periodo == 'hoje':
        return hoje, hoje, 'hoje', 'Hoje'
    if periodo == 'ontem':
        d = hoje - timedelta(days=1)
        return d, d, 'ontem', 'Ontem'
    if periodo == '7d':
        return hoje - timedelta(days=6), hoje, '7d', 'Últimos 7 dias'
    if periodo == 'mes':
        return hoje.replace(day=1), hoje, 'mes', 'Este mês'
    if periodo == 'personalizado':
        try:
            di = date.fromisoformat(data_inicio_raw)
            df = date.fromisoformat(data_fim_raw)
            label = f'{di.strftime("%d/%m/%Y")} a {df.strftime("%d/%m/%Y")}'
            return di, df, 'personalizado', label
        except (ValueError, TypeError):
            pass
    # default: 30d
    return hoje - timedelta(days=29), hoje, '30d', 'Últimos 30 dias'


# ── Dashboard de Marketing ─────────────────────────────────────────────────────

STATUS_PAGOS = ('pagamento_confirmado', 'em_separacao', 'pronto_retirada', 'enviado', 'entregue')


@staff_member_required
def dashboard_marketing(request):
    from apps.pedidos.models import Pedido

    data_inicio, data_fim, periodo, label_periodo = _parse_periodo(request)

    qs = Pedido.objects.filter(
        criado_em__date__gte=data_inicio,
        criado_em__date__lte=data_fim,
        status__in=STATUS_PAGOS,
    )

    # Resumo geral
    agg = qs.aggregate(
        total_pedidos=Count('id'),
        total_faturado=Sum('total'),
        ticket_medio=Avg('total'),
    )
    total_pedidos   = agg['total_pedidos'] or 0
    total_faturado  = agg['total_faturado'] or Decimal('0')
    ticket_medio    = agg['ticket_medio'] or Decimal('0')

    # Clientes únicos e novos x recorrentes
    clientes_ids = list(
        qs.filter(cliente__isnull=False).values_list('cliente_id', flat=True).distinct()
    )
    total_clientes = len(clientes_ids)
    novos = recorrentes = 0
    if clientes_ids:
        primeiros = dict(
            Pedido.objects.filter(status__in=STATUS_PAGOS, cliente_id__in=clientes_ids)
            .values('cliente_id')
            .annotate(primeiro=Min('criado_em'))
            .values_list('cliente_id', 'primeiro')
        )
        for dt in primeiros.values():
            if dt is None:
                continue
            try:
                d = timezone.localtime(dt).date()
            except Exception:
                d = dt if not hasattr(dt, 'date') else dt.date()
            if data_inicio <= d <= data_fim:
                novos += 1
            else:
                recorrentes += 1

    # Qualidade de atribuição
    com_utm  = qs.exclude(utm_source='').count()
    sem_utm  = total_pedidos - com_utm
    pct_attr = round(com_utm / total_pedidos * 100) if total_pedidos else 0

    # Ordenação de campanha
    ordenar_camp = request.GET.get('ordenar_campanha', 'receita')
    if ordenar_camp not in ('receita', 'pedidos', 'ticket'):
        ordenar_camp = 'receita'
    order_camp = {'receita': '-receita', 'pedidos': '-pedidos', 'ticket': '-ticket'}[ordenar_camp]

    def _enrich(rows, nome_key, vazio_label):
        for r in rows:
            r[nome_key] = r[nome_key] or vazio_label
            r['receita_fmt'] = _fmt(r.get('receita') or Decimal('0'))
            if 'ticket' in r:
                r['ticket_fmt'] = _fmt(r.get('ticket') or Decimal('0'))
            if total_faturado:
                r['pct'] = round(float(r.get('receita') or 0) / float(total_faturado) * 100)
            else:
                r['pct'] = 0
        return rows

    por_origem = _enrich(
        list(qs.values('utm_source').annotate(
            pedidos=Count('id'), receita=Sum('total'), ticket=Avg('total'),
        ).order_by('-receita')),
        'utm_source', '(direto / sem UTM)'
    )

    por_campanha = _enrich(
        list(qs.values('utm_campaign').annotate(
            pedidos=Count('id'), receita=Sum('total'), ticket=Avg('total'),
        ).order_by(order_camp)),
        'utm_campaign', '(sem campanha)'
    )

    por_midia = _enrich(
        list(qs.values('utm_medium').annotate(
            pedidos=Count('id'), receita=Sum('total'),
        ).order_by('-receita')),
        'utm_medium', '(sem mídia)'
    )

    # Pedidos recentes com dados de atribuição
    pedidos_recentes = list(
        qs.select_related('cliente').order_by('-criado_em').values(
            'numero', 'nome_completo', 'email', 'total',
            'utm_source', 'utm_campaign', 'criado_em', 'status',
        )[:15]
    )
    for p in pedidos_recentes:
        p['total_fmt']    = _fmt(p['total'] or Decimal('0'))
        p['utm_source']   = p['utm_source'] or '—'
        p['utm_campaign'] = p['utm_campaign'] or '—'
        try:
            p['criado_em_fmt'] = timezone.localtime(p['criado_em']).strftime('%d/%m %H:%M')
        except Exception:
            p['criado_em_fmt'] = str(p['criado_em'])[:16]

    context = {
        'title': 'Marketing e Atribuicao',
        'periodo': periodo,
        'label_periodo': label_periodo,
        'data_inicio_str': data_inicio.isoformat(),
        'data_fim_str': data_fim.isoformat(),
        # Resumo
        'total_pedidos':  total_pedidos,
        'total_faturado': _fmt(total_faturado),
        'ticket_medio':   _fmt(ticket_medio),
        'total_clientes': total_clientes,
        'novos':          novos,
        'recorrentes':    recorrentes,
        # Atribuição
        'com_utm':  com_utm,
        'sem_utm':  sem_utm,
        'pct_attr': pct_attr,
        # Tabelas
        'por_origem':    por_origem,
        'por_campanha':  por_campanha,
        'por_midia':     por_midia,
        'top_origens':   por_origem[:10],
        'top_campanhas': sorted(por_campanha, key=lambda r: -(r.get('receita') or 0))[:10],
        'pedidos_recentes': pedidos_recentes,
        'ordenar_campanha': ordenar_camp,
    }
    return render(request, 'admin/marketing_dashboard.html', context)


@staff_member_required
def dashboard_marketing_export(request):
    from apps.pedidos.models import Pedido
    import csv as csv_module
    import io

    data_inicio, data_fim, periodo, _ = _parse_periodo(request)

    qs = Pedido.objects.filter(
        criado_em__date__gte=data_inicio,
        criado_em__date__lte=data_fim,
        status__in=STATUS_PAGOS,
    )

    tipo    = request.GET.get('tipo', 'origem')
    formato = request.GET.get('formato', 'csv')

    if tipo == 'campanha':
        rows = list(qs.values('utm_campaign').annotate(
            pedidos=Count('id'), receita=Sum('total'), ticket=Avg('total'),
        ).order_by('-receita'))
        headers   = ['Campanha', 'Pedidos', 'Receita (R$)', 'Ticket Médio (R$)']
        data_rows = [
            [r['utm_campaign'] or '(sem campanha)', r['pedidos'],
             float(r['receita'] or 0), round(float(r['ticket'] or 0), 2)]
            for r in rows
        ]
        nome = 'campanhas'
    elif tipo == 'midia':
        rows = list(qs.values('utm_medium').annotate(
            pedidos=Count('id'), receita=Sum('total'),
        ).order_by('-receita'))
        headers   = ['Mídia', 'Pedidos', 'Receita (R$)']
        data_rows = [
            [r['utm_medium'] or '(sem mídia)', r['pedidos'], float(r['receita'] or 0)]
            for r in rows
        ]
        nome = 'midia'
    else:
        rows = list(qs.values('utm_source').annotate(
            pedidos=Count('id'), receita=Sum('total'), ticket=Avg('total'),
        ).order_by('-receita'))
        headers   = ['Origem', 'Pedidos', 'Receita (R$)', 'Ticket Médio (R$)']
        data_rows = [
            [r['utm_source'] or '(direto / sem UTM)', r['pedidos'],
             float(r['receita'] or 0), round(float(r['ticket'] or 0), 2)]
            for r in rows
        ]
        nome = 'origens'

    periodo_label = f'{data_inicio.strftime("%Y%m%d")}-{data_fim.strftime("%Y%m%d")}'
    filename = f'marketing_{nome}_{periodo_label}'

    if formato == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = nome.capitalize()

        # Header row com estilo
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = PatternFill('solid', fgColor='0A0A0A')
            cell.alignment = Alignment(horizontal='center')

        for row in data_rows:
            ws.append(row)

        # Ajusta largura das colunas
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    # CSV (padrão)
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = csv_module.writer(response)
    writer.writerow(headers)
    for row in data_rows:
        writer.writerow(row)
    return response

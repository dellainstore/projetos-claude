import csv as csv_module
import io
import time
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import openpyxl

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods

from apps.core.decorators import login_obrigatorio, papel_required, perm_required
from apps.produtos.services import precos as svc

BR_TZ = ZoneInfo("America/Sao_Paulo")


@login_obrigatorio
@perm_required("precos.ver")
def view_precos(request):
    modelos = svc.listar_modelos()
    modelo_selecionado = request.GET.get("modelo", "").strip()
    job_id = request.GET.get("job", "").strip()
    job = svc.get_job(job_id) if job_id else None
    return render(request, "produtos/precos.html", {
        "modelos": modelos,
        "modelo_selecionado": modelo_selecionado,
        "pode_exportar": request.user.tem_perm("precos.exportar_atacado"),
        "job": job,
    })


@login_obrigatorio
@papel_required("superadmin", "gestor")
def htmx_cores_por_modelo(request):
    base_name = request.GET.get("modelo", "").strip()
    if not base_name:
        return HttpResponse("")
    cores = svc.listar_cores_por_modelo(base_name)
    return render(request, "produtos/_precos_tabela.html", {
        "cores": cores,
        "base_name": base_name,
    })


@login_obrigatorio
@papel_required("superadmin", "gestor")
@require_http_methods(["POST"])
def view_aplicar_precos(request):
    base_name = request.POST.get("base_name", "").strip()
    color_keys = request.POST.getlist("cores")

    def _parse(field):
        val = request.POST.get(field, "").strip().replace(",", ".")
        try:
            v = float(val)
            return v if v > 0 else None
        except (ValueError, TypeError):
            return None

    varejo = _parse("varejo")
    custo = _parse("custo")
    atacado = _parse("atacado")

    if not base_name or not color_keys:
        messages.error(request, "Selecione um modelo e pelo menos uma cor.")
        return redirect("produtos:precos")

    if varejo is None and custo is None and atacado is None:
        messages.error(request, "Informe pelo menos um preço para atualizar.")
        return redirect(f"produtos:precos")

    job_id = svc.iniciar_job_precos(
        base_name=base_name,
        color_keys=color_keys,
        varejo=varejo,
        custo=custo,
        atacado=atacado,
        usuario=request.user.username,
    )

    return redirect(f"/produtos/precos/?modelo={base_name}&job={job_id}")


@login_obrigatorio
@papel_required("superadmin", "gestor")
def view_job_precos_status(request, job_id: str):
    """HTMX polling — retorna fragment com progresso do job."""
    job = svc.get_job(job_id)
    if not job:
        return HttpResponse('<div id="job-status-box"></div>')
    return render(request, "produtos/_precos_job_status.html", {"job": job})


@login_obrigatorio
@papel_required("superadmin", "gestor")
def view_download_csv_job(request, job_id: str):
    """Download do CSV de atacado gerado pelo job de background."""
    job = svc.get_job(job_id)
    if not job or not job.get("csv_atacado"):
        messages.warning(request, "Nenhum arquivo de atacado disponível para este job.")
        return redirect("produtos:precos")
    response = HttpResponse(job["csv_atacado"], content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="atacado_job.csv"'
    return response


@login_obrigatorio
@papel_required("superadmin", "gestor")
def view_download_csv_atacado(request):
    csv_data = request.session.pop("csv_atacado_pendente", None)
    nome = request.session.pop("csv_atacado_nome", "atacado.csv")
    if not csv_data:
        messages.warning(request, "Nenhum arquivo de atacado pendente.")
        return redirect("produtos:precos")
    response = HttpResponse(csv_data, content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{nome}"'
    return response


@login_obrigatorio
@papel_required("superadmin", "gestor")
def htmx_buscar_modelos_precos(request):
    from apps.produtos.services.business.lookup import search_base_names
    q = request.GET.get("q", "").strip()
    modelos = search_base_names(q, limit=40) if len(q) >= 2 else []
    return render(request, "produtos/_precos_modelos_options.html", {"modelos": modelos, "q": q})


@login_obrigatorio
@papel_required("superadmin", "gestor")
@require_http_methods(["POST"])
def view_upload_atacado(request):
    """Lê Excel com colunas SKU e Atacado; atualiza price_history e gera CSV para Bling."""
    arquivo = request.FILES.get("arquivo_atacado")
    if not arquivo:
        messages.error(request, "Nenhum arquivo enviado.")
        return redirect("produtos:precos")

    try:
        wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
        ws = wb.active

        header = [str(c.value or "").strip().upper() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_sku = None
        col_atac = None
        for i, h in enumerate(header):
            if "SKU" in h or "CODIGO" in h or "CÓDIGO" in h:
                col_sku = i
            if "ATAC" in h or "PRECO" in h or "PREÇO" in h or "VALOR" in h:
                col_atac = i

        if col_sku is None or col_atac is None:
            messages.error(request, f"Colunas SKU e Atacado não encontradas. Colunas detectadas: {', '.join(header)}")
            return redirect("produtos:precos")

        from apps.produtos.services.db import get_conn
        conn = get_conn()

        linhas = []
        skus_nao_encontrados = []
        ok = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            sku_raw = row[col_sku]
            atac_raw = row[col_atac]
            if not sku_raw or not atac_raw:
                continue

            sku = str(sku_raw).strip().upper()
            try:
                valor = float(str(atac_raw).replace(",", "."))
            except (ValueError, TypeError):
                continue

            variant = conn.execute(
                "SELECT bling_product_id, base_name, color_key FROM variants_cache WHERE UPPER(sku)=? AND active=1 LIMIT 1",
                (sku,)
            ).fetchone()

            if variant:
                import time
                conn.execute(
                    """INSERT INTO price_history (base_name, color_key, sku, bling_id, tipo, valor_antes, valor_novo, usuario, alterado_em)
                       VALUES (?, ?, ?, ?, 'atacado_local', NULL, ?, ?, ?)""",
                    (variant["base_name"], variant["color_key"], sku, variant["bling_product_id"],
                     valor, request.user.username, int(time.time()))
                )
                ok += 1
            else:
                skus_nao_encontrados.append(sku)

            linhas.append({"sku": sku, "valor": valor, "encontrado": variant is not None})

        conn.commit()
        conn.close()

        # Gera CSV dos não encontrados para importação manual no Bling
        if skus_nao_encontrados:
            buf = io.StringIO()
            w = csv_module.writer(buf)
            w.writerow(["Código", "Preço Atacado"])
            for l in linhas:
                if not l["encontrado"]:
                    w.writerow([l["sku"], f"{l['valor']:.2f}".replace(".", ",")])
            request.session["csv_atacado_pendente"] = buf.getvalue()
            request.session["csv_atacado_nome"] = "atacado_nao_encontrados.csv"
            messages.warning(
                request,
                f"{ok} preço(s) salvo(s) localmente. "
                f"{len(skus_nao_encontrados)} SKU(s) não encontrados no catálogo — "
                f"baixe o CSV para importar no Bling."
            )
        else:
            messages.success(request, f"{ok} preço(s) de atacado importados com sucesso.")

    except Exception as exc:
        messages.error(request, f"Erro ao processar o arquivo: {exc}")

    return redirect("produtos:precos")


@login_obrigatorio
@perm_required("precos.exportar_atacado")
def view_exportar_atacado_csv(request):
    """Exporta TODOS os preços de atacado (mais recente por SKU) como XLSX para o Bling."""
    from apps.produtos.services.db import get_conn
    conn = get_conn()
    rows = conn.execute("""
        SELECT ph.sku, ph.valor_novo, ph.base_name, ph.color_key,
               COALESCE(vc.product_name, ph.base_name) as product_name
        FROM price_history ph
        LEFT JOIN variants_cache vc ON UPPER(vc.sku) = UPPER(ph.sku) AND vc.active = 1
        INNER JOIN (
            SELECT UPPER(sku) as sku_upper, MAX(alterado_em) as max_ts
            FROM price_history
            WHERE tipo IN ('atacado', 'atacado_local')
              AND sku IS NOT NULL AND sku != ''
            GROUP BY UPPER(sku)
        ) latest ON UPPER(ph.sku) = latest.sku_upper AND ph.alterado_em = latest.max_ts
        WHERE ph.tipo IN ('atacado', 'atacado_local')
        ORDER BY ph.base_name, ph.color_key
    """).fetchall()

    # Marca todos como exportados
    now_ts = int(time.time())
    conn.execute("""
        UPDATE price_history
        SET exportado_em = ?
        WHERE tipo IN ('atacado', 'atacado_local')
          AND exportado_em IS NULL
    """, [now_ts])
    conn.commit()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atacado"
    ws.append(["Nome lista", "SKU", "Nome Produto", "Preço"])
    for row in rows:
        ws.append(["ATACADO", row["sku"], row["product_name"], round(float(row["valor_novo"]), 2)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="atacado_todos.xlsx"'
    return response


@login_obrigatorio
@perm_required("precos.exportar_atacado")
def view_exportar_atacado_ultimas(request):
    """Exporta apenas preços de atacado ainda não exportados (exportado_em IS NULL)."""
    from apps.produtos.services.db import get_conn
    conn = get_conn()

    # SKUs com ao menos um registro não exportado
    skus_raw = conn.execute("""
        SELECT DISTINCT UPPER(sku) as su
        FROM price_history
        WHERE tipo IN ('atacado', 'atacado_local')
          AND exportado_em IS NULL
          AND sku IS NOT NULL AND sku != ''
    """).fetchall()

    if not skus_raw:
        messages.info(request, "Nenhuma alteração pendente de exportação.")
        return redirect("produtos:precos")

    sku_uppers = [r["su"] for r in skus_raw]
    placeholders = ",".join("?" * len(sku_uppers))

    # Preço mais recente por SKU (de toda a history, não só não exportado)
    rows = conn.execute(f"""
        SELECT ph.sku, ph.valor_novo, ph.base_name, ph.color_key
        FROM price_history ph
        INNER JOIN (
            SELECT UPPER(sku) as sku_upper, MAX(alterado_em) as max_ts
            FROM price_history
            WHERE tipo IN ('atacado', 'atacado_local')
              AND UPPER(sku) IN ({placeholders})
            GROUP BY UPPER(sku)
        ) latest ON UPPER(ph.sku) = latest.sku_upper AND ph.alterado_em = latest.max_ts
        WHERE ph.tipo IN ('atacado', 'atacado_local')
        ORDER BY ph.base_name, ph.color_key
    """, sku_uppers).fetchall()

    # Marca todos os registros não exportados desses SKUs como exportados
    now_ts = int(time.time())
    conn.execute(f"""
        UPDATE price_history
        SET exportado_em = ?
        WHERE tipo IN ('atacado', 'atacado_local')
          AND exportado_em IS NULL
          AND UPPER(sku) IN ({placeholders})
    """, [now_ts, *sku_uppers])
    conn.commit()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atacado"
    ws.append(["Nome lista", "SKU", "Nome Produto", "Preço"])

    # Busca nome do produto para cada SKU
    conn2 = get_conn()
    for row in rows:
        vc = conn2.execute(
            "SELECT product_name FROM variants_cache WHERE UPPER(sku)=? AND active=1 LIMIT 1",
            (row["sku"].upper(),)
        ).fetchone()
        product_name = vc["product_name"] if vc and vc["product_name"] else row["base_name"]
        ws.append(["ATACADO", row["sku"], product_name, round(float(row["valor_novo"]), 2)])
    conn2.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="atacado_ultimas.xlsx"'
    return response


@login_obrigatorio
@perm_required("precos.alterar")
@require_http_methods(["POST"])
def view_excluir_preco(request, preco_id: int):
    from apps.produtos.services.db import get_conn
    conn = get_conn()
    deleted = conn.execute("DELETE FROM price_history WHERE id = ?", (preco_id,)).rowcount
    conn.commit()
    conn.close()
    if deleted:
        messages.success(request, f"Registro #{preco_id} removido.")
    else:
        messages.warning(request, f"Registro #{preco_id} não encontrado.")
    return redirect(request.POST.get("next") or "produtos:historico_precos")


@login_obrigatorio
@perm_required("precos.ver")
def view_historico_precos(request):
    hoje_br = datetime.now(tz=BR_TZ).date()
    periodo = request.GET.get("periodo", "hoje")
    data_inicio_str = request.GET.get("data_inicio", "")
    data_fim_str = request.GET.get("data_fim", "")

    if periodo == "hoje":
        start = datetime(hoje_br.year, hoje_br.month, hoje_br.day, tzinfo=BR_TZ)
        end = start + timedelta(days=1)
    elif periodo == "ontem":
        ontem = hoje_br - timedelta(days=1)
        start = datetime(ontem.year, ontem.month, ontem.day, tzinfo=BR_TZ)
        end = start + timedelta(days=1)
    elif periodo == "7dias":
        end = datetime(hoje_br.year, hoje_br.month, hoje_br.day, tzinfo=BR_TZ) + timedelta(days=1)
        start = end - timedelta(days=7)
    elif periodo == "30dias":
        end = datetime(hoje_br.year, hoje_br.month, hoje_br.day, tzinfo=BR_TZ) + timedelta(days=1)
        start = end - timedelta(days=30)
    elif periodo == "personalizado":
        try:
            start = datetime.strptime(data_inicio_str, "%Y-%m-%d").replace(tzinfo=BR_TZ)
        except (ValueError, TypeError):
            start = datetime(hoje_br.year, hoje_br.month, 1, tzinfo=BR_TZ)
        try:
            end = datetime.strptime(data_fim_str, "%Y-%m-%d").replace(tzinfo=BR_TZ) + timedelta(days=1)
        except (ValueError, TypeError):
            end = datetime(hoje_br.year, hoje_br.month, hoje_br.day, tzinfo=BR_TZ) + timedelta(days=1)
    else:
        end = datetime(hoje_br.year, hoje_br.month, hoje_br.day, tzinfo=BR_TZ) + timedelta(days=1)
        start = end - timedelta(days=7)

    pagina = max(1, int(request.GET.get("pagina", 1)))
    filtros = {
        "base_name": request.GET.get("modelo", ""),
        "color_key": request.GET.get("cor", ""),
        "tipo": request.GET.get("tipo", ""),
        "periodo": periodo,
        "data_inicio": data_inicio_str,
        "data_fim": data_fim_str,
    }

    hist = svc.historico_precos(
        pagina=pagina,
        por_pagina=50,
        base_name=filtros["base_name"],
        color_key=filtros["color_key"],
        tipo=filtros["tipo"],
        start_ts=int(start.timestamp()),
        end_ts=int(end.timestamp()),
    )

    # Converte timestamps para datetime em horário de Brasília
    for r in hist["registros"]:
        ts = r.get("alterado_em")
        r["alterado_em_dt"] = datetime.fromtimestamp(ts, tz=BR_TZ) if ts else None

    periodos = [
        ("hoje", "Hoje"),
        ("ontem", "Ontem"),
        ("7dias", "7 dias"),
        ("30dias", "30 dias"),
        ("personalizado", "Personalizado"),
    ]

    return render(request, "produtos/historico_precos.html", {
        "hist": hist,
        "filtros": filtros,
        "pagina": pagina,
        "periodos": periodos,
        "pode_exportar": request.user.tem_perm("precos.exportar_atacado"),
        "pode_alterar_preco": request.user.tem_perm("precos.alterar"),
    })
